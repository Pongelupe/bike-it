#!/usr/bin/env python3
from openpyxl import Workbook
import psycopg2
import sys

def get_query_coverage(cover = 0, region = -1, view = ''):
    where = 'MIN > %(r)s'
    
    if cover == 1:
        where = 'MIN <= %(r)s AND MAX > %(r)s'
    if cover == 2:
        where = f'MIN <= %(r)s AND MAX <= %(r)s'

    query =   f"""SELECT COUNT(*)
FROM (SELECT id,
             MIN,
             MAX
      FROM distance_per_segment_pampulha
      WHERE {where}
      UNION
      SELECT id,
             MIN,
             MAX
      FROM distance_per_segment_central
      WHERE {where}
      UNION
      SELECT id,
             MIN,
             MAX
      FROM distance_per_segment_orla_pampulha
      WHERE {where}) AS a""" if region == -1 else f"""SELECT COUNT(*)
FROM (SELECT id,
             MIN,
             MAX
      FROM {view}
      WHERE {where}) as a
"""
    return query


if __name__ == '__main__':
    region = sys.argv[1]
    dest_file = f'scatter_{region}.xlsx'
    view = 'distance_per_segment_central'
    if region == '1':
        view = 'distance_per_segment_pampulha'

    if region == '2':
        view = 'distance_per_segment_orla_pampulha'

    wb = Workbook()
    ws = wb.active
    ws['B1'] = 'Sem Cobertura'
    ws['C1'] = 'Parcialmente Cobertas'
    ws['D1'] ='Totalmente Cobertas'

    con = psycopg2.connect(host='localhost', database='bikeit',
        user='postgres', password='root')
    cursor = con.cursor()
    radius = [0, 30, 60, 90, 120, 150, 180, 210, 240, 270, 300, 330, 360, 390, 420, 450, 480, 510] 
    line = 2
    for r in radius:
        ws.cell(row = line, column = 1).value = r
        for cover in range(3):
            cursor.execute( get_query_coverage(cover, region, view), {"r": r})
            recset = cursor.fetchone()
            print(recset[0])
            ws.cell(row = line, column = cover + 2).value = recset[0]
        line += 1

    wb.save(filename = dest_file)
